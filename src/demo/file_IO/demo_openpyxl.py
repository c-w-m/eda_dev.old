"""
Link: Working with Excel sheets in Python using openpyxl
URL: https://medium.com/aubergine-solutions/working-with-excel-sheets-in-python-using-openpyxl-4f9fd32de87f
"""

from openpyxl import Workbook

# set demo path_file
path_file = r'data/fileIO_openpyxl_01.xlsx'

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# Create a new Workbook object
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# this creates a new worksheet (just not yet saved)
wb = Workbook()

# get the name of all sheets - deprecated
print('1. initial -> print(wb.get_sheet_names()) - deprecated')
print(wb.get_sheet_names())

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# Create a new sheet called 'Data1' and 'Data2'
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# create a new worksheet, you can specify the index, and arrange its position in the workbook as appropriate
# is scheduled to the second worksheet, index=0 is the first position
wb.create_sheet('Data1', index=1)
wb.create_sheet('Data2', index=2)
wb.create_sheet('Junk1', index=3)

# get the name of all sheets - deprecated
print('2. added sheets -> print(wb.get_sheet_names()) - deprecated')
print(wb.get_sheet_names())

# save to file
wb.save(path_file)

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# Delete the 'Sheet' and  'Junk1' sheets
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# direct assignment can change the name of the worksheet
sheet = wb.get_sheet_by_name('Junk1')
# delete a worksheet
wb.remove(sheet)

sheet = wb.get_sheet_by_name('Sheet')
# delete a worksheet
wb.remove(sheet)

# get the name of all sheets - deprecated
print('3. deleted sheets -> print(wb.get_sheet_names()) - deprecated')
print(wb.get_sheet_names())

# save to file
wb.save(path_file)

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# Write a `row` to `Data1`
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# add a line to 'Data'
row = [1, 2, 3, 4, 5]
sheet = wb.get_sheet_by_name('Data1')
sheet.append(row)

# save to file
wb.save(path_file)

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# Write a `rows` to `Data2`
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# add multiple lines to 'Sheet1'
data = [
    ('Number', 'data1', 'data2'),
    (2, 20, 200),
    (3, 30, 300),
    (4, 40, 400),
    (5, 50, 500),
    (6, 60, 600),
]

sheet = wb.get_sheet_by_name('Data2')
# append all rows
for row in data:
    sheet.append(row)

cell = sheet.cell(row=2, column=1)
print('sheet.cell(row=2, column=1).value = {}'.format(cell.value))

cell.value = '=SUM(B2:C2)'
print('sheet.cell(row=2, column=1).value = {}'.format(cell.value))

# save to file
wb.save(path_file)
